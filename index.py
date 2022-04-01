import sys
from shutil import copyfile

from openpyxl import load_workbook


def process(source, target):
    # 1.打开 Excel 表格并获取表格名称
    source_workbook = load_workbook(filename=source)
    # 2.通过 sheet 名称获取表格
    source_sheet = source_workbook["1-1"]
    people = []
    print("Source文件有这么多行：" + str(source_sheet.max_row))
    for index in range(4, source_sheet.max_row):
        name = source_sheet['H' + str(index)].value
        number = source_sheet['I' + str(index)].value
        if name and number:
            people.append(People(name, number))
            print(name)
    print(len(people))

    count = int(len(people) / 499)

    target_files = []
    for index in range(1, count + 2):
        target_file = target.replace('.xlsx', '-%s.xlsx' % index)
        target_files.append(target_file)
        copyfile(target, target_file)

    n = 499
    people_group = [people[i:i + n] for i in range(0, len(people), n)]
    print(len(people_group))
    for group in people_group:
        file_path = target_files.pop(0)
        print("处理文件中：" + file_path)
        target_workbook = load_workbook(filename=file_path)
        target_sheet = target_workbook["导入模板"]
        index = 3
        for person in group:
            print(person.name)
            target_sheet['A' + str(index)] = person.name
            target_sheet['C' + str(index)] = person.number
            index += 1
        target_workbook.save(file_path)

    print('end')


class People:
    def __init__(self, name, number):
        self.name = name
        self.number = number


if __name__ == '__main__':
    print('参数列表:', str(sys.argv))
    if len(sys.argv) < 3:
        print('请传入源文件和目标文件路径')
        exit()
    print(sys.argv[1])
    print(sys.argv[2])
    process(sys.argv[1], sys.argv[2])
